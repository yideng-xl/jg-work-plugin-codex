import openpyxl
import pytest
from src.summary_io import write_summary, read_summary, write_travel_detail, HEADERS

ROWS = [
    {"seq": 1, "filename": "a.pdf", "seller": "曹操出行", "invoice_kind": "普票",
     "copies": 1, "date": "2026-05-01", "amount": 185.82, "amount_usd": None,
     "tax": None, "category": "差旅费-交通费", "category_uncertain": False,
     "from_": "A", "to": "B", "loc_uncertain": False, "reason": "无直达交通",
     "note": "", "is_prepaid": False, "is_allowance": False, "allowance_days": None},
    {"seq": 2, "filename": "claude.pdf", "seller": "Anthropic", "invoice_kind": "普票",
     "copies": 1, "date": "2026-05-02", "amount": None, "amount_usd": 82.35,
     "tax": None, "category": "待确认", "category_uncertain": True,
     "from_": None, "to": None, "loc_uncertain": False, "reason": "",
     "note": "美元票", "is_prepaid": False, "is_allowance": False, "allowance_days": None},
]

def test_write_summary(tmp_path):
    out = tmp_path / "摘要.xlsx"
    write_summary(ROWS, str(out), rate=6.78)
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws["A1"].value == "汇率"
    assert ws["B1"].value == 6.78
    # 第 3 行表头，含"费用分类科目"
    headers = [c.value for c in ws[3]]
    assert "费用分类科目" in headers
    # 美元票金额列写成公式
    amount_col = headers.index("金额") + 1
    usd_cell = ws.cell(row=5, column=amount_col)   # 第 2 张票在第 5 行
    assert str(usd_cell.value).startswith("=") and "82.35" in str(usd_cell.value)
    # 标黄：第 2 张票科目单元格有填充
    cat_col = headers.index("费用分类科目") + 1
    assert ws.cell(row=5, column=cat_col).fill.fgColor.rgb[-6:].upper() == "FFFF00"


def test_usd_rows_require_explicit_rate(tmp_path):
    out = tmp_path / "摘要.xlsx"
    with pytest.raises(ValueError, match="必须传入已确认的汇率"):
        write_summary(ROWS, str(out))


def _base_row(**overrides):
    row = {"seq": 1, "filename": "a.pdf", "seller": "曹操出行", "invoice_kind": "普票",
           "copies": 1, "date": "2026-05-01", "amount": 100.0, "amount_usd": None,
           "tax": None, "category": "差旅费-交通费", "category_uncertain": False,
           "from_": "A", "to": "B", "loc_uncertain": False, "reason": "无直达交通",
           "note": "", "is_prepaid": False, "is_allowance": False, "allowance_days": None}
    row.update(overrides)
    return row


def test_zhuanpiao_tax_only_for_zhuanpiao(tmp_path):
    rows = [
        _base_row(seq=1, invoice_kind="普票", tax=5.41),
        _base_row(seq=2, invoice_kind="专票", tax=10.0),
    ]
    out = tmp_path / "摘要.xlsx"
    write_summary(rows, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    headers = [c.value for c in ws[3]]
    tax_col = headers.index("专票税额") + 1
    # 普票行（第 4 行）：即使携带 tax 值，专票税额列也必须为空
    assert ws.cell(row=4, column=tax_col).value is None
    # 专票行（第 5 行）：专票税额列应显示 tax 值
    assert ws.cell(row=5, column=tax_col).value == 10.0


def test_prepaid_row_fully_red_filled(tmp_path):
    rows = [_base_row(is_prepaid=True)]
    out = tmp_path / "摘要.xlsx"
    write_summary(rows, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=4, column=col)
        assert cell.fill.fgColor.rgb[-6:].upper() == "FFC7CE", f"col {col} not red-filled"


def test_loc_uncertain_yellow_fill(tmp_path):
    rows = [_base_row(loc_uncertain=True)]
    out = tmp_path / "摘要.xlsx"
    write_summary(rows, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    headers = [c.value for c in ws[3]]
    from_col = headers.index("出发地") + 1
    to_col = headers.index("目的地") + 1
    assert ws.cell(row=4, column=from_col).fill.fgColor.rgb[-6:].upper() == "FFFF00"
    assert ws.cell(row=4, column=to_col).fill.fgColor.rgb[-6:].upper() == "FFFF00"


def test_clean_row_no_fill(tmp_path):
    rows = [_base_row(category_uncertain=False, loc_uncertain=False, is_prepaid=False)]
    out = tmp_path / "摘要.xlsx"
    write_summary(rows, str(out))
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    headers = [c.value for c in ws[3]]
    cat_col = headers.index("费用分类科目") + 1
    cell = ws.cell(row=4, column=cat_col)
    # openpyxl 默认无填充时 patternType 为 None（未写入任何 solid fill）
    assert cell.fill.patternType is None


def test_read_summary_roundtrip(tmp_path):
    out = tmp_path / "摘要.xlsx"
    write_summary(ROWS, str(out), rate=6.78)
    back = read_summary(str(out))
    assert back[0]["seller"] == "曹操出行"
    assert back[0]["category"] == "差旅费-交通费"
    assert back[0]["date"] == "2026-05-01"


def test_read_summary_usd_amount_not_formula_string(tmp_path):
    # 回归锁：金额列对美元票写的是 Excel 公式（=82.35*$B$1），read_summary
    # 必须以 data_only 打开，读回来的是缓存的数值（或未在 Excel 中打开过时为
    # None），绝不能是公式字符串——否则阶段二 Claude 填 OA 会拿到一坨公式文本，
    # 且若该行落入交通费明细，write_travel_detail 里的 float(...) 会直接报错。
    out = tmp_path / "摘要.xlsx"
    write_summary(ROWS, str(out), rate=6.78)
    back = read_summary(str(out))

    rmb_amount = back[0]["amount"]
    usd_amount = back[1]["amount"]

    # RMB 行是字面量，原样读回数字
    assert rmb_amount == 185.82

    # USD 行是公式：绝不能读成以 "=" 开头的字符串
    assert not (isinstance(usd_amount, str) and usd_amount.startswith("="))
    # 允许的两种安全结果：None（文件从未在 Excel 中打开保存过，无缓存值）
    # 或者一个数字（缓存值）
    assert usd_amount is None or isinstance(usd_amount, (int, float))


def test_write_travel_detail(tmp_path):
    rows = [
        {"date": "2026-04-14", "from_": "酒店", "to": "郑州办公室",
         "amount": 12.9, "reason": "无直达交通"},
        {"date": "2026-04-15", "from_": "郑州办公室", "to": "郑州东站",
         "amount": 24.29, "reason": "无直达交通"},
    ]
    out = tmp_path / "交通费明细.xlsx"
    write_travel_detail(rows, str(out), applicant="张三")
    import openpyxl
    ws = openpyxl.load_workbook(str(out)).active
    assert ws.title == "交通费明细"
    assert ws["A1"].value == "上海巨耕-交通费报销明细"
    assert [c.value for c in ws[2]][:5] == ["日期", "出发地", "目的地", "金额", "打车原因"]
    # 数据落位：日期/出发地/目的地/金额/打车原因
    assert ws["A3"].value == "2026-04-14"
    assert ws["B3"].value == "酒店"
    assert ws["D4"].value == 24.29
    # 合计行紧接数据（2 行数据 → 第 5 行），金额列是 SUM 公式（照模板，非预算数）
    assert ws["A5"].value == "合计"
    assert ws["D5"].value == "=SUM(D3:D4)"
    # 报销人行在合计下空一行（第 7 行），无边框
    assert ws["D7"].value == "报销人："
    assert ws["E7"].value == "张三"
    assert ws["D7"].border.bottom.style is None

    # 模板样式保真：宋体、字号、边框、合并、列宽
    assert ws["A1"].font.name == "宋体" and ws["A1"].font.size == 22 and ws["A1"].font.bold
    assert "A1:E1" in [str(m) for m in ws.merged_cells.ranges]
    assert ws["A2"].font.size == 16              # 表头
    assert ws["A3"].font.size == 11              # 数据
    assert ws["A5"].font.size == 16              # 合计标签
    assert ws["A3"].border.left.style == "thin"  # 数据有网格线
    assert ws["A5"].border.top.style == "thin"   # 合计有网格线
    assert round(ws.column_dimensions["B"].width, 2) == 22.34
    assert round(ws.row_dimensions[1].height, 2) == 51.75


def test_write_travel_detail_requires_applicant(tmp_path):
    out = tmp_path / "交通费明细.xlsx"
    with pytest.raises(ValueError, match="实际报销人"):
        write_travel_detail([], str(out))
