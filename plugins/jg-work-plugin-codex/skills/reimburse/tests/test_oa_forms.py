import openpyxl
from src.oa_forms import (write_oa_general_detail, write_oa_travel_detail,
                          infer_trip_defaults, default_header,
                          aggregate_by_category, _inject_totals)


def _rows_text(ws):
    return [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]


def test_aggregate_by_category():
    items = [
        {"category": "会议费", "date": "2026-05-01", "amount": 100.0,
         "tax": None, "note": "a"},
        {"category": "会议费", "date": "2026-05-03", "amount": 106.0,
         "tax": 6.0, "note": "b"},
        {"category": "差旅费-交通费", "legs": "1", "amount": 50.0,
         "tax": None, "note": "c"},
        {"category": "差旅费-交通费", "legs": "2", "amount": 30.0,
         "tax": None, "note": "d"},
    ]
    agg = aggregate_by_category(items)
    assert len(agg) == 2                    # 两类 → 两行
    meeting = agg[0]
    assert meeting["category"] == "会议费"
    assert meeting["amount"] == 206.0       # 合计金额
    assert meeting["tax"] == 6.0            # 合计税额
    assert meeting["date"] == "2026-05-01~2026-05-03"   # 日期跨度
    assert meeting["note"] is None                       # 备注一律不填
    transport = agg[1]
    assert transport["legs"] == "12"        # 覆盖行程 1、2 的并集
    assert transport["tax"] == 0.0          # 普票即使含税也写 0（不写空）


def test_general_detail_aggregates(tmp_path):
    # 同类多票 → 聚合成一类一行（OA 报销费用明细口径）
    items = [
        {"category": "市内交通费", "date": "2026-05-01", "amount": 100.0,
         "tax": None, "note": ""},
        {"category": "市内交通费", "date": "2026-05-02", "amount": 106.0,
         "tax": 6.0, "note": ""},
    ]
    out = tmp_path / "报销金额明细.xlsx"
    write_oa_general_detail(items, str(out))
    ws = openpyxl.load_workbook(str(out)).active
    assert ws["A1"].value == "上海巨耕-报销金额明细"
    assert [c.value for c in ws[2]] == ["序号", "费用分类", "费用发生日期",
                                        "报销金额", "增值税专票税额", "费用金额", "备注"]
    # 两票同类 → 只 1 行数据（第 3 行），合计在第 4 行
    assert ws["B3"].value == "市内交通费"
    assert ws["D3"].value == 206.0          # 报销金额合计
    assert ws["F3"].value == 200.0          # 费用金额 = 206 - 6
    assert ws["A4"].value == "合计"
    assert ws["D4"].value == "=SUM(D3:D3)"
    assert ws["A3"].font.name == "宋体"
    assert ws["A3"].border.left.style == "thin"


def test_general_detail_with_header_block(tmp_path):
    items = [{"category": "会议费", "date": "2026-05-01", "amount": 100.0,
              "tax": None, "note": ""}]
    fields = default_header("通用", reason="claude 订阅", applicant="张三",
                            fill_date="2026-07-09")
    out = tmp_path / "带表头.xlsx"
    write_oa_general_detail(items, str(out), header_fields=fields)
    ws = openpyxl.load_workbook(str(out)).active
    # 表头区把明细往下推：第 1 行大标题，之后是表头信息区
    flat = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    assert "申请人" in flat and "张三" in flat
    assert "报销类型" in flat and "通用" in flat
    assert "claude 订阅" in flat
    # 明细表头仍然存在（被推到更下面）
    assert any([ws.cell(row=r, column=c).value for c in range(1, 8)][:2]
               == ["序号", "费用分类"] for r in range(1, ws.max_row + 1))


def test_default_header_requires_applicant():
    import pytest
    with pytest.raises(ValueError, match="实际申请人"):
        default_header("通用")


def test_travel_detail_two_tables(tmp_path):
    trips = [
        {"dep_date": "2026-06-09", "dep_time": "08:12", "from_": "上海虹桥",
         "arr_date": "2026-06-09", "arr_time": "10:00", "to": "宁波",
         "stay_days": 1},
    ]
    items = [
        {"category": "差旅费-交通费", "legs": "1", "amount": 151.0,
         "tax": None, "note": "高铁"},
        {"category": "差旅补贴", "legs": "1", "amount": 50.0,
         "tax": None, "note": "1天"},
    ]
    out = tmp_path / "差旅报销明细.xlsx"
    write_oa_travel_detail(trips, items, str(out))
    ws = openpyxl.load_workbook(str(out)).active
    assert ws["A1"].value == "上海巨耕-差旅报销明细"
    band_texts = _rows_text(ws)
    assert "行程明细" in band_texts
    assert "报销费用明细" in band_texts
    # 行程列出现在报销费用明细表头
    exp_head_row = band_texts.index("报销费用明细") + 2
    assert [c.value for c in ws[exp_head_row]][:7] == [
        "序号", "费用分类", "行程", "报销金额", "增值税专票税额", "费用金额", "费用备注"]
    assert any(r[0] == "合计" for r in ws.iter_rows(values_only=True))


def test_infer_trip_defaults():
    trips = [
        {"dep_date": "2026-06-09", "dep_time": "08:12", "from_": "上海虹桥",
         "to": "宁波"},
        {"dep_date": "2026-06-10", "dep_time": "18:30", "from_": "宁波",
         "to": "上海虹桥"},
    ]
    out = infer_trip_defaults(trips)
    # 到达日期补成出发日期（不标黄）
    assert out[0]["arr_date"] == "2026-06-09"
    assert "arr_date" not in out[0]["_uncertain"]
    # 到达时间按 +2h 推、标黄
    assert out[0]["arr_time"] == "10:12"
    assert "arr_time" in out[0]["_uncertain"]
    # 住宿天数逐段：首段 6/9→6/10 = 1 夜；末段返程 = 0；都标黄
    assert out[0]["stay_days"] == 1
    assert "stay_days" in out[0]["_uncertain"]
    assert out[1]["stay_days"] == 0
    # 不改原对象
    assert "arr_time" not in trips[0]


def test_train_and_flight_arrival_time_requires_public_schedule():
    trips = [
        {"transport_type": "高铁", "dep_date": "2099-01-10", "dep_time": "08:00"},
        {"transport_type": "飞机", "dep_date": "2099-01-12", "dep_time": "16:00"},
    ]
    out = infer_trip_defaults(trips)
    assert out[0].get("arr_time") is None
    assert out[1].get("arr_time") is None
    assert "arr_time" in out[0]["_uncertain"]
    assert "arr_time" in out[1]["_uncertain"]


def _flatten(rows):
    return {label: v for pairs in rows for (label, v) in pairs}


def test_inject_totals_fills_header():
    items = [{"amount": 981.65, "tax": None}, {"amount": 577.12, "tax": 32.67},
             {"amount": 125.0, "tax": 0.0}]
    rows = default_header("差旅", reason="郑州集中办公", applicant="张三")
    filled = _flatten(_inject_totals(rows, items))
    assert filled["报销总金额"] == 1683.77       # Σ报销金额
    assert filled["增值税专票税额合计"] == 32.67   # Σ专票税额
    assert filled["费用合计"] == 1651.10          # 报销总金额 - 专票税额合计


def test_lodging_special_invoice_keeps_tax_in_reimbursement(tmp_path):
    """专票价税合计进入报销金额，只有费用金额扣税。"""
    items = [{"category": "差旅-住宿费", "legs": "1", "amount": 500.0,
              "tax": 50.0, "note": ""}]
    rows = default_header("差旅", reason="测试出差", applicant="测试用户")
    filled = _flatten(_inject_totals(rows, items))
    assert filled["报销总金额"] == 500.0
    assert filled["增值税专票税额合计"] == 50.0
    assert filled["费用合计"] == 450.0

    out = tmp_path / "住宿专票口径.xlsx"
    write_oa_travel_detail([], items, str(out), header_fields=rows)
    ws = openpyxl.load_workbook(str(out), data_only=False).active
    band_texts = _rows_text(ws)
    first_expense = band_texts.index("报销费用明细") + 3
    assert ws.cell(first_expense, 4).value == 500.0
    assert ws.cell(first_expense, 5).value == 50.0
    assert ws.cell(first_expense, 6).value == 450.0


def test_infer_uncertain_cells_highlighted_yellow(tmp_path):
    trips = infer_trip_defaults([
        {"dep_date": "2026-06-09", "dep_time": "08:12", "from_": "上海虹桥",
         "to": "宁波"},
        {"dep_date": "2026-06-10", "dep_time": "18:30", "from_": "宁波",
         "to": "上海虹桥"},
    ])
    out = tmp_path / "t.xlsx"
    write_oa_travel_detail(trips, [], str(out))
    ws = openpyxl.load_workbook(str(out)).active
    band_texts = _rows_text(ws)
    band_row = band_texts.index("行程明细") + 1   # 转 1-based 行号
    first_trip = band_row + 2                      # band → 表头 → 首数据行
    # 到达时间在第 6 列（F），推断值应标黄
    assert ws.cell(row=first_trip, column=6).fill.fgColor.rgb.endswith("FFFF00")


def test_default_header_travel_has_extra_fields():
    rows = default_header("差旅", reason="宁波出差", applicant="张三")
    flat = _flatten(rows)
    assert "出差天数" in flat and "实际工作天数" in flat
    assert "出差申请记录" in flat and "附件" in flat
    assert flat["报销类型"] == "差旅"
    assert flat["费用大区"] == "公共"           # 固定值预填
    # 报销事由 整行铺满（独占一行、单组）
    assert any(len(p) == 1 and p[0][0] == "报销事由" for p in rows)


def test_default_header_general_has_attachment_fullwidth():
    rows = default_header("通用", reason="抵票", applicant="张三")
    assert any(len(p) == 1 and p[0][0] == "附件" for p in rows)   # 通用附件整行
    flat = _flatten(rows)
    assert "出差天数" not in flat               # 通用无差旅字段
