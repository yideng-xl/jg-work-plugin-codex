import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

YELLOW = PatternFill("solid", fgColor="FFFF00")
RED = PatternFill("solid", fgColor="FFC7CE")

# 交通费报销明细表标准模板的固定样式（照 90-invoice/模板/交通费报销明细表.xlsx
# 逐属性抽出，"文字、线都要一样"）。改这里前先比对模板，别凭印象调。
_THIN = Side(style="thin")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER = Alignment(horizontal="center", vertical="center")
_HCENTER = Alignment(horizontal="center")
_VCENTER = Alignment(vertical="center")
_TD_COL_WIDTHS = {"A": 11.78, "B": 22.34, "C": 15.0, "D": 8.22, "E": 24.66}
_TD_HEADERS = ["日期", "出发地", "目的地", "金额", "打车原因"]

HEADERS = ["序号", "文件名", "卖方/平台", "票种", "打印份数", "费用发生日期",
           "金额", "专票税额", "费用分类科目", "出发地", "目的地",
           "打车原因", "备注"]


def write_summary(rows: list, out_path: str, rate: float | None = None):
    if any(r.get("amount_usd") is not None for r in rows) and rate is None:
        raise ValueError("存在美元票，必须传入已确认的汇率")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报销摘要"

    ws["A1"] = "汇率"
    ws["B1"] = rate

    for col, h in enumerate(HEADERS, start=1):
        ws.cell(row=3, column=col, value=h).font = Font(bold=True)

    col_of = {h: i + 1 for i, h in enumerate(HEADERS)}
    for i, r in enumerate(rows):
        row = 4 + i
        ws.cell(row=row, column=col_of["序号"], value=r["seq"])
        ws.cell(row=row, column=col_of["文件名"], value=r["filename"])
        ws.cell(row=row, column=col_of["卖方/平台"], value=r["seller"])
        ws.cell(row=row, column=col_of["票种"], value=r["invoice_kind"])
        ws.cell(row=row, column=col_of["打印份数"], value=r["copies"])
        ws.cell(row=row, column=col_of["费用发生日期"], value=r["date"])

        amt_cell = ws.cell(row=row, column=col_of["金额"])
        if r.get("amount_usd") is not None:
            amt_cell.value = f"={r['amount_usd']}*$B$1"
        else:
            amt_cell.value = r.get("amount")

        tax_value = r.get("tax") if r.get("invoice_kind") == "专票" else None
        ws.cell(row=row, column=col_of["专票税额"], value=tax_value)
        cat_cell = ws.cell(row=row, column=col_of["费用分类科目"], value=r.get("category"))
        from_cell = ws.cell(row=row, column=col_of["出发地"], value=r.get("from_"))
        to_cell = ws.cell(row=row, column=col_of["目的地"], value=r.get("to"))
        ws.cell(row=row, column=col_of["打车原因"], value=r.get("reason") or "")
        ws.cell(row=row, column=col_of["备注"], value=r.get("note") or "")

        if r.get("category_uncertain"):
            cat_cell.fill = YELLOW
        if r.get("loc_uncertain"):
            from_cell.fill = YELLOW
            to_cell.fill = YELLOW
        if r.get("is_prepaid"):
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=row, column=col).fill = RED

    wb.save(out_path)


def read_summary(path: str) -> list:
    # data_only=True：openpyxl 从不计算公式，只能读取 Excel 缓存的计算结果。
    # 金额列对美元票写的是公式（=82.35*$B$1，见 write_summary），必须用
    # data_only 打开才能读回换算后的人民币数值，而不是公式字符串本身。
    # 若该文件从未在 Excel 中打开并保存过（没有缓存值），公式单元格会读成
    # None——这也是安全的：float(None or 0) == 0；一旦有缓存值就是正常数字。
    # 唯一不允许出现的是形如 "=82.35*$B$1" 的公式字符串。
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[3]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[idx["序号"]] is None:
            continue
        rows.append({
            "seq": row[idx["序号"]],
            "filename": row[idx["文件名"]],
            "seller": row[idx["卖方/平台"]],
            "invoice_kind": row[idx["票种"]],
            "copies": row[idx["打印份数"]],
            "date": row[idx["费用发生日期"]],
            "amount": row[idx["金额"]],
            "tax": row[idx["专票税额"]],
            "category": row[idx["费用分类科目"]],
            "from_": row[idx["出发地"]],
            "to": row[idx["目的地"]],
            "reason": row[idx["打车原因"]],
            "note": row[idx["备注"]],
        })
    return rows


def write_travel_detail(rows: list, out_path: str):
    """严格照抄 90-invoice/模板/交通费报销明细表.xlsx：宋体、边框、列宽、
    行高、合计公式、报销人无边框，都跟模板一致。data 行数按传入 rows 动态。
    列映射：日期←date / 出发地←from_ / 目的地←to / 金额←amount / 打车原因←reason。
    过滤（交通类、剔预充值）由调用方做，本函数原样写入。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "交通费明细"

    n = len(rows)
    first_data = 3
    total_row = first_data + n          # 合计行
    signer_row = total_row + 2          # 报销人行（模板里合计与报销人间隔一空行）

    # 列宽
    for col, w in _TD_COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    # 行高：标题 51.75，表头到合计 20.1
    ws.row_dimensions[1].height = 51.75
    for r in range(2, total_row + 1):
        ws.row_dimensions[r].height = 20.1

    # 网格边框：标题行到合计行，每个单元格四边 thin（标题为合并单元格，
    # 内部竖线不会渲染，只显示外框，与模板一致）
    for r in range(1, total_row + 1):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = _BORDER_ALL

    # 标题（合并 A1:E1，宋体 22 加粗，居中）
    ws.merge_cells("A1:E1")
    a1 = ws["A1"]
    a1.value = "上海巨耕-交通费报销明细"
    a1.font = Font(name="宋体", size=22, bold=True)
    a1.alignment = _CENTER

    # 表头（宋体 16，水平居中）
    for col, h in enumerate(_TD_HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(name="宋体", size=16)
        cell.alignment = _HCENTER

    # 数据行（宋体 11；日期/出发地/目的地居中，金额/打车原因默认左对齐，同模板）
    for i, item in enumerate(rows):
        r = first_data + i
        vals = [item.get("date"), item.get("from_"), item.get("to"),
                item.get("amount"), item.get("reason")]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = Font(name="宋体", size=11)
            if col <= 3:
                cell.alignment = _CENTER

    # 合计行（A 列"合计"宋体 16 居中；D 列 =SUM 公式）
    for c in range(1, 6):
        ws.cell(row=total_row, column=c).font = Font(name="宋体", size=11)
    a_total = ws.cell(row=total_row, column=1, value="合计")
    a_total.font = Font(name="宋体", size=16)
    a_total.alignment = _HCENTER
    if n:
        ws.cell(row=total_row, column=4,
                value=f"=SUM(D{first_data}:D{total_row - 1})")

    # 报销人行（无边框，宋体 11，垂直居中）
    d_sign = ws.cell(row=signer_row, column=4, value="报销人：")
    e_sign = ws.cell(row=signer_row, column=5, value="许磊")
    for cell in (d_sign, e_sign):
        cell.font = Font(name="宋体", size=11)
        cell.alignment = _VCENTER

    wb.save(out_path)
