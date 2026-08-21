"""泛微 OA「技术报销」表单的报销金额明细，做成 Excel 方便照着填 OA、也给
二期插件当数据源。列结构照 90-invoice/模板/ 下 通用报销.png / 差旅报销.png 两张
截图，样式沿用交通费明细（宋体 + thin 网格线 + 合并标题）。

两版：
- write_oa_general_detail：通用报销，表头信息区 + 一张「报销金额明细」表。
- write_oa_travel_detail：差旅报销，表头信息区 + 两张表（「行程明细」+
  「报销费用明细」，费用明细多一个「行程」列，值是行程明细表的序号组合，
  如 123 覆盖行程 1/2/3）。

费用金额 = 报销金额 - 增值税专票税额（专票才有税额，普票留空按 0 算）。

**重要：OA 报销费用明细是「一个费用分类一行」**，不是一票一行——同类发票金额
合计成一行（明细逐票在交通费明细表/A4 发票拼贴那边）。这跟交通费明细表（一票/
一段行程一行）是两码事，别弄混。写表前先 aggregate_by_category 聚合。

推断默认值（用户口径：能推的先推、标黄让用户核对，别留空）：
- 到达日期默认 = 出发日期（火车/城际多数当天到）。
- 到达时间推不出准值，按车种给保守默认并标黄。
- 住宿天数 = 往返日期跨度的夜数，落在首行、标黄。
这些由 infer_trip_defaults 生成，写表时 _uncertain 里的列标黄。
"""
from collections import OrderedDict

import openpyxl
from openpyxl.styles import Font, Alignment

from src.summary_io import _THIN, _BORDER_ALL, _CENTER, _HCENTER, YELLOW

_LEFT = Alignment(horizontal="left", vertical="center")

# 当前模板的组织字段默认值。申请人不在这里固定，必须由本次报销信息传入。
_FIXED = {"费用大区": "公共", "费用部门": "产品部", "费用地区": "总部"}

_F_TITLE = Font(name="宋体", size=18, bold=True)
_F_BAND = Font(name="宋体", size=14, bold=True)
_F_HEAD = Font(name="宋体", size=11, bold=True)
_F_BODY = Font(name="宋体", size=11)
_F_LABEL = Font(name="宋体", size=11, bold=True)

_GEN_COLS = ["序号", "费用分类", "费用发生日期", "报销金额",
             "增值税专票税额", "费用金额", "备注"]
_TRIP_COLS = ["序号", "出发日期", "出发时间", "出发地",
              "到达日期", "到达时间", "目的地", "住宿天数"]
# 行程明细列 key，跟 _TRIP_COLS 对应，供 _uncertain 标黄用
_TRIP_KEYS = ["_seq", "dep_date", "dep_time", "from_",
              "arr_date", "arr_time", "to", "stay_days"]
_EXP_COLS = ["序号", "费用分类", "行程", "报销金额",
             "增值税专票税额", "费用金额", "费用备注"]   # 差旅用"费用备注"（通用用"备注"）

# 到达时间估算的默认在途时长（小时），推不准，仅作占位、写表时标黄
_DEFAULT_LEG_HOURS = 2


def _col_letter(i):
    return openpyxl.utils.get_column_letter(i)


def _band(ws, row, ncols, text):
    """整行合并的分区标题条（如"行程明细""报销费用明细"）。"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _F_BAND
    c.alignment = _CENTER
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = _BORDER_ALL


def _header(ws, row, cols, ncols=None):
    for i, h in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _F_HEAD
        c.alignment = _HCENTER
        c.border = _BORDER_ALL
    # 表宽比列数大时（差旅费用明细 7 列铺在 8 列宽上），补齐右侧占位边框
    for col in range(len(cols) + 1, (ncols or len(cols)) + 1):
        ws.cell(row=row, column=col).border = _BORDER_ALL


def _title(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = _F_TITLE
    c.alignment = _CENTER
    ws.row_dimensions[row].height = 40
    for col in range(1, ncols + 1):
        ws.cell(row=row, column=col).border = _BORDER_ALL


def _write_header_block(ws, start_row, ncols, rows):
    """表头信息区，照真实 OA 单双列布局。rows 是「行」列表，每行是 1~2 组
    (标签, 值)：两组 → 各占半宽；一组 → 值整行铺满（如 报销事由/附件）。
    值为 None/"" 的标黄提示手填。返回块结束后的下一行。"""
    half = ncols // 2
    r = start_row
    for pairs in rows:
        if len(pairs) == 1:                  # 整行铺满：标签 1 列，值占 2..ncols
            label, val = pairs[0]
            lab = ws.cell(row=r, column=1, value=label)
            lab.font, lab.alignment, lab.border = _F_LABEL, _HCENTER, _BORDER_ALL
            if ncols >= 2:
                ws.merge_cells(start_row=r, start_column=2,
                               end_row=r, end_column=ncols)
            vcell = ws.cell(row=r, column=2, value=val)
            vcell.font, vcell.alignment = _F_BODY, _LEFT
            for col in range(1, ncols + 1):
                ws.cell(row=r, column=col).border = _BORDER_ALL
            if val in (None, ""):
                vcell.fill = YELLOW
        else:                                # 双列：标签+半宽值 ×2
            for j, (label, val) in enumerate(pairs):
                base = 1 + j * half
                lab = ws.cell(row=r, column=base, value=label)
                lab.font, lab.alignment, lab.border = _F_LABEL, _HCENTER, _BORDER_ALL
                vstart, vend = base + 1, base + half - 1
                if vend >= vstart:
                    ws.merge_cells(start_row=r, start_column=vstart,
                                   end_row=r, end_column=vend)
                vcell = ws.cell(row=r, column=vstart, value=val)
                vcell.font, vcell.alignment = _F_BODY, _CENTER
                for col in range(base, base + half):
                    ws.cell(row=r, column=col).border = _BORDER_ALL
                if val in (None, ""):
                    vcell.fill = YELLOW
            for col in range(2 * half + 1, ncols + 1):   # 尾列补边框
                ws.cell(row=r, column=col).border = _BORDER_ALL
        r += 1
    return r


def _fee_amount(amount, tax):
    return round(float(amount or 0) - float(tax or 0), 2)


def aggregate_by_category(items: list) -> list:
    """把逐票 items 按费用分类聚合成「一类一行」（OA 报销费用明细的口径）。
    每项 in：category/amount/tax/date(通用)/legs(差旅)/note。
    每项 out：category/amount(合计)/tax(合计,全无税则None)/date(日期跨度 min~max)/
             legs(该类覆盖的行程序号并集,如"12")/note(多票时"共N笔")/count。
    保持首次出现顺序。"""
    groups = OrderedDict()
    for it in items:
        cat = it.get("category")
        g = groups.get(cat)
        if g is None:
            g = {"category": cat, "amount": 0.0, "tax": 0.0, "has_tax": False,
                 "dates": [], "legs": set(), "count": 0}
            groups[cat] = g
        g["amount"] += float(it.get("amount") or 0)
        if it.get("tax") is not None:
            g["tax"] += float(it["tax"])
            g["has_tax"] = True
        if it.get("date"):
            g["dates"].append(str(it["date"]))
        for ch in str(it.get("legs") or ""):
            if ch.isdigit():
                g["legs"].add(ch)
        g["count"] += 1

    out = []
    for g in groups.values():
        ds = sorted(g["dates"])
        date = None
        if ds:
            date = ds[0] if ds[0] == ds[-1] else f"{ds[0]}~{ds[-1]}"
        legs = "".join(sorted(g["legs"])) or None
        # 税额规则（用户口径）：专票写真实税额，普票即使含税也写 0——所以聚合后
        # 税额恒为数字（0.0 表示该类全是普票/无专票税），不写空。费用金额=报销金额−税额。
        # 备注列一律不填（省麻烦），需要时用户自己在 Excel 里加。
        out.append({"category": g["category"], "amount": round(g["amount"], 2),
                    "tax": round(g["tax"], 2),
                    "date": date, "legs": legs, "note": None,
                    "count": g["count"]})
    return out


def _fill_body(cell, value, uncertain=False, center=True):
    cell.value = value
    cell.font = _F_BODY
    cell.border = _BORDER_ALL
    if center:
        cell.alignment = _CENTER
    if uncertain:
        cell.fill = YELLOW


def _sum_row(ws, row, first, n, ncols, label_col, sum_cols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.font = _F_BODY
        c.border = _BORDER_ALL
    lab = ws.cell(row=row, column=label_col, value="合计")
    lab.font = _F_LABEL
    lab.alignment = _HCENTER
    if n:
        for col in sum_cols:
            L = _col_letter(col)
            ws.cell(row=row, column=col,
                    value=f"=SUM({L}{first}:{L}{row - 1})").alignment = _CENTER


def _apply_widths(ws, widths):
    for i, w in widths.items():
        ws.column_dimensions[_col_letter(i)].width = w


def write_oa_general_detail(items: list, out_path: str, header_fields=None,
                            aggregate=True):
    """通用报销：表头信息区（可选）+ 报销金额明细表（一个费用分类一行）。
    items 逐票传入（category/date/amount/tax/note），默认按费用分类聚合成一类一行。
    费用金额自动 = 报销金额 - 专票税额；合计用 SUM 公式。"""
    if aggregate:
        items = aggregate_by_category(items)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报销金额明细"
    ncols = len(_GEN_COLS)

    _title(ws, 1, ncols, "上海巨耕-报销金额明细")
    row = 2
    if header_fields:
        header_fields = _inject_totals(header_fields, items)
        row = _write_header_block(ws, row, ncols, header_fields)
        row += 1                             # 表头区与明细之间空一行

    _header(ws, row, _GEN_COLS)
    first = row + 1
    for i, it in enumerate(items):
        r = first + i
        vals = [i + 1, it.get("category"), it.get("date"), it.get("amount"),
                it.get("tax"), _fee_amount(it.get("amount"), it.get("tax")),
                None]                            # 备注列一律不填
        for col, v in enumerate(vals, start=1):
            _fill_body(ws.cell(row=r, column=col), v)

    total = first + len(items)
    _sum_row(ws, total, first, len(items), ncols, label_col=1, sum_cols=(4, 5, 6))
    _apply_widths(ws, {1: 6, 2: 18, 3: 15, 4: 12, 5: 15, 6: 12, 7: 24})
    wb.save(out_path)


def write_oa_travel_detail(trips: list, items: list, out_path: str,
                           header_fields=None, aggregate=True):
    """差旅报销：表头信息区（可选）+ 行程明细 + 报销费用明细两张表。
    trips 每项（逐段行程，一段一行）：dep_date/dep_time/from_/arr_date/arr_time/
      to/stay_days（可带 _uncertain=集合，成员是这些 key，标黄）。
    items 逐票传入（category/legs/amount/tax/note，legs 是行程序号组合字符串），
      报销费用明细默认按费用分类聚合成一类一行，legs 取该类覆盖的行程序号并集。"""
    if aggregate:
        items = aggregate_by_category(items)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "差旅报销明细"
    ncols = len(_TRIP_COLS)                  # 8 列，取较宽的做整体基准

    _title(ws, 1, ncols, "上海巨耕-差旅报销明细")
    row = 2
    if header_fields:
        header_fields = _inject_totals(header_fields, items)
        row = _write_header_block(ws, row, ncols, header_fields)
        row += 1

    # 表一：行程明细
    _band(ws, row, ncols, "行程明细")
    _header(ws, row + 1, _TRIP_COLS)
    first_trip = row + 2
    for i, tp in enumerate(trips):
        r = first_trip + i
        unc = tp.get("_uncertain") or set()
        vals = [tp.get("_seq", i + 1), tp.get("dep_date"), tp.get("dep_time"),
                tp.get("from_"), tp.get("arr_date"), tp.get("arr_time"),
                tp.get("to"), tp.get("stay_days")]
        for col, (key, v) in enumerate(zip(_TRIP_KEYS, vals), start=1):
            _fill_body(ws.cell(row=r, column=col), v, uncertain=key in unc)

    # 表二：报销费用明细（空一行分隔）
    band2 = first_trip + len(trips) + 1
    _band(ws, band2, ncols, "报销费用明细")
    exp_head = band2 + 1
    _header(ws, exp_head, _EXP_COLS, ncols=ncols)
    first_exp = exp_head + 1
    for i, it in enumerate(items):
        r = first_exp + i
        vals = [i + 1, it.get("category"), it.get("legs"),
                it.get("amount"), it.get("tax"),
                _fee_amount(it.get("amount"), it.get("tax")), None]  # 备注不填
        for col, v in enumerate(vals, start=1):
            _fill_body(ws.cell(row=r, column=col), v)
        for col in range(len(_EXP_COLS) + 1, ncols + 1):
            ws.cell(row=r, column=col).border = _BORDER_ALL

    total = first_exp + len(items)
    _sum_row(ws, total, first_exp, len(items), ncols, label_col=1, sum_cols=(4, 5, 6))
    _apply_widths(ws, {1: 6, 2: 14, 3: 12, 4: 14, 5: 12, 6: 12, 7: 14, 8: 10})
    wb.save(out_path)


def _add_hours(hhmm, hours):
    """"08:12" + 2h -> "10:12"；解析失败返回 None。"""
    try:
        h, m = str(hhmm).split(":")
        h = (int(h) + hours) % 24
        return f"{h:02d}:{int(m):02d}"
    except (ValueError, AttributeError):
        return None


def _nights_between(d1, d2):
    """两个 'YYYY-MM-DD' 之间的夜数；解析失败返回 None。"""
    from datetime import date
    try:
        parts = [tuple(int(x) for x in str(d).split("-")) for d in (d1, d2)]
        return (date(*parts[1]) - date(*parts[0])).days
    except (ValueError, TypeError):
        return None


def infer_trip_defaults(trips: list) -> list:
    """给行程补推断默认值并标黄。返回新 trips 列表，每项带 _uncertain 集合，
    原地不改传入对象。规则（对齐真实 OA 差旅单）：
    - 到达日期缺 → = 出发日期（当天到，通常成立，不标黄）。
    - 高铁/飞机到达时间缺 → 保持空值，必须先查公开计划时刻。
    - 其他交通到达时间缺 → 出发时间 +默认时长（推不准，标黄）。
    - 住宿天数逐段：本段到达 → 下段出发之间的夜数；末段（返程）为 0；标黄。
      例：4/13 到郑州、4/15 离郑州 → 首段住宿 2、末段 0。"""
    out = []
    for tp in trips:
        t = dict(tp)
        unc = set(t.get("_uncertain") or set())
        if not t.get("arr_date") and t.get("dep_date"):
            t["arr_date"] = t["dep_date"]          # 当天到，通常成立，不标黄
        public_schedule_required = t.get("transport_type") in {"高铁", "飞机"}
        if not t.get("arr_time") and t.get("dep_time") and not public_schedule_required:
            guess = _add_hours(t["dep_time"], _DEFAULT_LEG_HOURS)
            if guess:
                t["arr_time"] = guess
                unc.add("arr_time")
        if not t.get("arr_time") and public_schedule_required:
            unc.add("arr_time")
        t["_uncertain"] = unc
        out.append(t)

    # 住宿天数逐段：本段到达日 → 下段出发日的夜数；最后一段为 0
    for i, t in enumerate(out):
        if t.get("stay_days") is not None:
            continue
        if i + 1 < len(out):
            nights = _nights_between(t.get("arr_date") or t.get("dep_date"),
                                     out[i + 1].get("dep_date"))
            t["stay_days"] = nights if nights is not None else None
        else:
            t["stay_days"] = 0                     # 末段返程，无住宿
        if t["stay_days"] is not None:
            t["_uncertain"].add("stay_days")
    return out


# 表头里这三个合计由明细自动算（对齐真实 OA 差旅单的 费用合计/专票税额合计/报销总金额）
_AUTO_TOTALS = {"报销总金额": "reimburse", "增值税专票税额合计": "tax",
                "费用合计": "fee"}


def _inject_totals(header_rows, items):
    """把表头里空着的 报销总金额/增值税专票税额合计/费用合计 用明细合计填上。
    报销总金额=Σ报销金额，专票税额合计=Σ专票税额，费用合计=报销总金额−专票税额合计。
    header_rows 是「行」列表（每行 1~2 组 (标签,值)）。"""
    if not header_rows:
        return header_rows
    reimburse = round(sum(float(it.get("amount") or 0) for it in items), 2)
    tax = round(sum(float(it.get("tax") or 0) for it in items), 2)
    vals = {"reimburse": reimburse, "tax": tax, "fee": round(reimburse - tax, 2)}
    out = []
    for pairs in header_rows:
        new_pairs = []
        for label, v in pairs:
            if label in _AUTO_TOTALS and (v is None or v == ""):
                v = vals[_AUTO_TOTALS[label]]
            new_pairs.append((label, v))
        out.append(new_pairs)
    return out


def _half_day_value(label: str, value) -> float:
    """OA 天数字段只接受非负的 0.5 天倍数。"""
    if value is None or value == "":
        raise ValueError(f"差旅报销必须填写{label}")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label}必须是数字") from exc
    if number < 0 or abs(number * 2 - round(number * 2)) > 1e-9:
        raise ValueError(f"{label}必须是非负的 0.5 天倍数")
    return number


def default_header(kind: str, reason: str = "", applicant: str | None = None,
                   fill_date: str = "", title: str = "", doc_no: str = "",
                   travel_days=None, actual_work_days=None) -> list:
    """OA 表头，照真实「技术报销」表单的字段顺序与双列布局。kind: "通用"/"差旅"。
    返回「行」列表：每行 1~2 组 (标签, 值)；单组行整行铺满（报销事由/通用的附件）。
    固定值（费用大区/部门/地区）预填，报销总金额/专票税额合计/费用合计 写表时自动算，
    其余留空写表时标黄。"""
    if not applicant or not applicant.strip():
        raise ValueError("必须传入本次报销的实际申请人")

    f = _FIXED
    rows = [
        [("标题", title), ("报销单号", doc_no)],
        [("申请人", applicant.strip()), ("填报日期", fill_date)],
        [("费用大区", f["费用大区"]), ("费用部门", f["费用部门"])],
        [("报销类型", kind), ("费用地区", f["费用地区"])],
        [("报销分类", ""), ("销售合同", "")],
        [("报销事由", reason)],                       # 整行铺满
    ]
    if kind == "差旅":
        travel_days = _half_day_value("出差天数", travel_days)
        actual_work_days = _half_day_value("实际工作天数", actual_work_days)
        if actual_work_days > travel_days:
            raise ValueError("实际工作天数不能大于出差天数")
        rows += [
            [("出差天数", travel_days), ("实际工作天数", actual_work_days)],
            [("费用合计", ""), ("增值税专票税额合计", "")],
            [("报销总金额", ""), ("出差申请记录", "")],
            [("报销付款日期", ""), ("附件", "")],
        ]
    else:
        rows += [
            [("附件", "")],                           # 通用：附件整行铺满
            [("费用合计", ""), ("增值税专票税额合计", "")],
            [("报销总金额", ""), ("报销付款日期", "")],
        ]
    return rows
